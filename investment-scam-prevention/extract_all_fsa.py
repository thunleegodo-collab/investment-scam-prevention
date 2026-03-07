"""
金融庁 登録業者一覧 Excel → JSON 変換スクリプト（3種類対応版）

対象ファイル:
  kinyushohin.xlsx  金融商品取引業者
  chuukai.xlsx      金融商品仲介業者
  touroku.xlsx      登録金融機関

使い方:
  1. pip install openpyxl
  2. 各Excelを同フォルダに置く
  3. python extract_all_fsa.py
  4. fsa_all.json が生成される
"""

import json
import re
import sys
from pathlib import Path
from datetime import datetime

sys.stdout.reconfigure(encoding="utf-8")

try:
    import openpyxl
except ImportError:
    print("openpyxl が見つかりません: pip install openpyxl")
    sys.exit(1)

OUTPUT_PATH = "fsa_all.json"
DATA_START_ROW = 8  # 全ファイル共通（8行目からデータ開始）


def normalize(text: str) -> str:
    if not text:
        return ""
    text = str(text)
    text = text.translate(str.maketrans(
        "ａｂｃｄｅｆｇｈｉｊｋｌｍｎｏｐｑｒｓｔｕｖｗｘｙｚ"
        "ＡＢＣＤＥＦＧＨＩＪＫＬＭＮＯＰＱＲＳＴＵＶＷＸＹＺ"
        "０１２３４５６７８９",
        "abcdefghijklmnopqrstuvwxyz"
        "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        "0123456789"
    ))
    text = re.sub(
        r'株式会社|有限会社|合同会社|合資会社|合名会社|一般社団法人|一般財団法人'
        r'|\(株\)|\(有\)|（株）|（有）', '', text
    )
    text = re.sub(r'[\s\u3000]+', '', text)
    return text.lower()


def cell_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip().replace("\n", " / ").replace("\r", "")


def excel_date(value) -> str:
    """Excelシリアル値 or datetime を文字列に変換"""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)) and value > 1000:
        # Excel シリアル値をdatetimeに変換（1900年基準）
        from datetime import timedelta
        base = datetime(1899, 12, 30)
        return (base + timedelta(days=int(value))).strftime("%Y-%m-%d")
    return cell_str(value)


# ============================================================
# 金融商品取引業者 (kinyushohin.xlsx)
# ============================================================
def extract_kinyushohin(xlsx_path: str) -> list[dict]:
    COL_REG_NO   = 1
    COL_REG_DATE = 2
    COL_NAME     = 3
    COL_ADDRESS  = 6
    COL_PHONE    = 7
    COL_TYPE1    = 8
    COL_TYPE2    = 9
    COL_ADVISORY = 10
    COL_MGMT     = 11

    companies = []
    seen = set()
    print(f"読み込み中: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    for i, row in enumerate(ws.iter_rows(min_row=DATA_START_ROW, values_only=True), DATA_START_ROW):
        name = cell_str(row[COL_NAME] if len(row) > COL_NAME else None)
        if not name:
            continue
        key = normalize(name)
        if not key or key in seen:
            continue
        seen.add(key)

        companies.append({
            "name":     name,
            "name_n":   key,
            "address":  cell_str(row[COL_ADDRESS]  if len(row) > COL_ADDRESS  else None),
            "addr_n":   normalize(cell_str(row[COL_ADDRESS] if len(row) > COL_ADDRESS else None)),
            "reg_no":   cell_str(row[COL_REG_NO]   if len(row) > COL_REG_NO   else None),
            "reg_date": excel_date(row[COL_REG_DATE] if len(row) > COL_REG_DATE else None),
            "phone":    cell_str(row[COL_PHONE]    if len(row) > COL_PHONE    else None),
            "type1":    cell_str(row[COL_TYPE1]    if len(row) > COL_TYPE1    else None),
            "type2":    cell_str(row[COL_TYPE2]    if len(row) > COL_TYPE2    else None),
            "advisory": cell_str(row[COL_ADVISORY] if len(row) > COL_ADVISORY else None),
            "mgmt":     cell_str(row[COL_MGMT]     if len(row) > COL_MGMT     else None),
            "category": "金融商品取引業者",
        })

    wb.close()
    print(f"  → {len(companies)} 件")
    return companies


# ============================================================
# 金融商品仲介業者 (chuukai.xlsx)
# ============================================================
def extract_chuukai(xlsx_path: str) -> list[dict]:
    COL_REG_NO   = 1
    COL_REG_DATE = 2
    COL_NAME     = 3
    COL_ADDRESS  = 6
    COL_PHONE    = 7
    COL_CORPTYPE = 8   # 法人又は個人の別
    COL_BELONGS  = 9   # 所属金融商品取引業者等

    companies = []
    seen = set()
    print(f"読み込み中: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    for i, row in enumerate(ws.iter_rows(min_row=DATA_START_ROW, values_only=True), DATA_START_ROW):
        name = cell_str(row[COL_NAME] if len(row) > COL_NAME else None)
        if not name:
            continue
        key = normalize(name)
        if not key or key in seen:
            continue
        seen.add(key)

        companies.append({
            "name":     name,
            "name_n":   key,
            "address":  cell_str(row[COL_ADDRESS]  if len(row) > COL_ADDRESS  else None),
            "addr_n":   normalize(cell_str(row[COL_ADDRESS] if len(row) > COL_ADDRESS else None)),
            "reg_no":   cell_str(row[COL_REG_NO]   if len(row) > COL_REG_NO   else None),
            "reg_date": excel_date(row[COL_REG_DATE] if len(row) > COL_REG_DATE else None),
            "phone":    cell_str(row[COL_PHONE]     if len(row) > COL_PHONE    else None),
            "corp_type": cell_str(row[COL_CORPTYPE] if len(row) > COL_CORPTYPE else None),
            "belongs":  cell_str(row[COL_BELONGS]   if len(row) > COL_BELONGS  else None),
            "category": "金融商品仲介業者",
        })

    wb.close()
    print(f"  → {len(companies)} 件")
    return companies


# ============================================================
# 登録金融機関 (touroku.xlsx)
# ============================================================
def extract_touroku(xlsx_path: str) -> list[dict]:
    COL_REG_NO   = 1
    COL_REG_DATE = 2
    COL_NAME     = 3
    COL_ADDRESS  = 6
    COL_PHONE    = 7

    companies = []
    seen = set()
    print(f"読み込み中: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    for i, row in enumerate(ws.iter_rows(min_row=DATA_START_ROW, values_only=True), DATA_START_ROW):
        name = cell_str(row[COL_NAME] if len(row) > COL_NAME else None)
        if not name:
            continue
        key = normalize(name)
        if not key or key in seen:
            continue
        seen.add(key)

        companies.append({
            "name":     name,
            "name_n":   key,
            "address":  cell_str(row[COL_ADDRESS]  if len(row) > COL_ADDRESS  else None),
            "addr_n":   normalize(cell_str(row[COL_ADDRESS] if len(row) > COL_ADDRESS else None)),
            "reg_no":   cell_str(row[COL_REG_NO]   if len(row) > COL_REG_NO   else None),
            "reg_date": excel_date(row[COL_REG_DATE] if len(row) > COL_REG_DATE else None),
            "phone":    cell_str(row[COL_PHONE]     if len(row) > COL_PHONE    else None),
            "category": "登録金融機関",
        })

    wb.close()
    print(f"  → {len(companies)} 件")
    return companies


def main():
    results = {}

    for key, fname, extractor in [
        ("kinyushohin", "kinyushohin.xlsx", extract_kinyushohin),
        ("chuukai",     "chuukai.xlsx",     extract_chuukai),
        ("touroku",     "touroku.xlsx",     extract_touroku),
    ]:
        p = Path(fname)
        if not p.exists():
            print(f"スキップ（ファイルなし）: {fname}")
            results[key] = []
        else:
            results[key] = extractor(str(p))

    all_companies = (
        results["kinyushohin"] +
        results["chuukai"] +
        results["touroku"]
    )

    output = {
        "generated":  datetime.now().strftime("%Y-%m-%d"),
        "count":      len(all_companies),
        "kinyushohin_count": len(results["kinyushohin"]),
        "chuukai_count":     len(results["chuukai"]),
        "touroku_count":     len(results["touroku"]),
        "kinyushohin": results["kinyushohin"],
        "chuukai":     results["chuukai"],
        "touroku":     results["touroku"],
    }

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, separators=(',', ':'))

    print(f"\n完了: {OUTPUT_PATH} に保存（計 {len(all_companies)} 件）")
    print(f"  金融商品取引業者: {len(results['kinyushohin'])} 件")
    print(f"  金融商品仲介業者: {len(results['chuukai'])} 件")
    print(f"  登録金融機関:     {len(results['touroku'])} 件")


if __name__ == "__main__":
    main()
