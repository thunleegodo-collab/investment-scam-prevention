"""
金融庁 金融商品取引業者登録一覧 Excel → JSON 変換スクリプト

使い方:
  1. pip install openpyxl
  2. 下記URLからExcelをダウンロード（ファイル名は kinyushohin.xlsx のまま）
     https://www.fsa.go.jp/menkyo/menkyoj/kinyushohin.xlsx
  3. このスクリプトと同じフォルダに xlsx を置く
  4. python extract_fsa_data.py
  5. fsa_advisors.json が生成される
  6. checker.html と同じフォルダに fsa_advisors.json を置く

注意:
  - 金融庁はExcelを定期更新します。最新版に差し替えたら再実行してください。
"""

import json
import re
import sys
from pathlib import Path
from datetime import datetime

sys.stdout.reconfigure(encoding="utf-8")

try:
    import openpyxl
except ImportError:
    print("openpyxl が見つかりません。以下を実行してください:")
    print("  pip install openpyxl")
    sys.exit(1)

XLSX_PATH   = "kinyushohin.xlsx"
OUTPUT_PATH = "fsa_advisors.json"

# Excelの列インデックス（0始まり）
COL_JURISDICTION = 0   # 所管
COL_REG_NO       = 1   # 登録番号
COL_REG_DATE     = 2   # 登録年月日
COL_NAME         = 3   # 金融商品取引業者名
COL_CORP_NO      = 4   # 法人番号
COL_POSTAL       = 5   # 郵便番号
COL_ADDRESS      = 6   # 本店等所在地
COL_PHONE        = 7   # 代表等電話番号
COL_TYPE1        = 8   # 第一種
COL_TYPE2        = 9   # 第二種
COL_ADVISORY     = 10  # 投資助言・代理業
COL_MANAGEMENT   = 11  # 投資運用業

DATA_START_ROW = 8  # データ開始行（1始まり）


def normalize(text: str) -> str:
    """会社名・住所を比較しやすい形に正規化する"""
    if not text:
        return ""
    text = str(text)
    # 全角英数を半角に変換
    text = text.translate(str.maketrans(
        "ａｂｃｄｅｆｇｈｉｊｋｌｍｎｏｐｑｒｓｔｕｖｗｘｙｚ"
        "ＡＢＣＤＥＦＧＨＩＪＫＬＭＮＯＰＱＲＳＴＵＶＷＸＹＺ"
        "０１２３４５６７８９",
        "abcdefghijklmnopqrstuvwxyz"
        "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        "0123456789"
    ))
    # 法人格表記を除去
    text = re.sub(r'株式会社|有限会社|合同会社|合資会社|合名会社|一般社団法人|一般財団法人'
                  r'|\(株\)|\(有\)|（株）|（有）', '', text)
    # スペース除去
    text = re.sub(r'[\s\u3000]+', '', text)
    return text.lower()


def cell_str(value) -> str:
    """セル値を文字列に変換"""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip().replace("\n", " ")


def extract_from_xlsx(xlsx_path: str) -> list[dict]:
    companies = []
    seen = set()

    print(f"Excelを読み込み中: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    total = ws.max_row

    for i, row in enumerate(ws.iter_rows(min_row=DATA_START_ROW, values_only=True), DATA_START_ROW):
        if i % 100 == 0:
            print(f"  行 {i}/{total} 処理中...", end="\r")

        name = cell_str(row[COL_NAME] if len(row) > COL_NAME else None)
        if not name:
            continue

        reg_no  = cell_str(row[COL_REG_NO]   if len(row) > COL_REG_NO   else None)
        address = cell_str(row[COL_ADDRESS]   if len(row) > COL_ADDRESS  else None)
        phone   = cell_str(row[COL_PHONE]     if len(row) > COL_PHONE    else None)
        reg_date = cell_str(row[COL_REG_DATE] if len(row) > COL_REG_DATE else None)
        type1    = cell_str(row[COL_TYPE1]    if len(row) > COL_TYPE1    else None)
        type2    = cell_str(row[COL_TYPE2]    if len(row) > COL_TYPE2    else None)
        advisory = cell_str(row[COL_ADVISORY] if len(row) > COL_ADVISORY else None)
        mgmt     = cell_str(row[COL_MANAGEMENT] if len(row) > COL_MANAGEMENT else None)

        key = normalize(name)
        if not key or key in seen:
            continue
        seen.add(key)

        companies.append({
            "name":     name,
            "name_n":   normalize(name),
            "address":  address,
            "addr_n":   normalize(address),
            "reg_no":   reg_no,
            "phone":    phone,
            "reg_date": reg_date,
            "type1":    type1,
            "type2":    type2,
            "advisory": advisory,
            "mgmt":     mgmt,
        })

    wb.close()
    print(f"\n合計 {len(companies)} 件の業者情報を抽出しました。")
    return companies


def main():
    xlsx_path = Path(XLSX_PATH)
    if not xlsx_path.exists():
        print(f"エラー: {XLSX_PATH} が見つかりません。")
        print("https://www.fsa.go.jp/menkyo/menkyoj/kinyushohin.xlsx からダウンロードしてください。")
        sys.exit(1)

    companies = extract_from_xlsx(str(xlsx_path))

    if not companies:
        print("警告: 業者情報が抽出できませんでした。Excelの形式を確認してください。")
        sys.exit(1)

    output = {
        "generated":  datetime.now().strftime("%Y-%m-%d"),
        "source":     "金融庁 金融商品取引業者登録一覧",
        "source_url": "https://www.fsa.go.jp/menkyo/menkyoj/kinyushohin.xlsx",
        "count":      len(companies),
        "companies":  companies,
    }

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"完了: {OUTPUT_PATH} に保存しました（{len(companies)} 件）")


if __name__ == "__main__":
    main()
